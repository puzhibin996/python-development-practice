from openpyxl import load_workbook

workbook = load_workbook(filename='test.xlsx')
print(workbook.sheetnames)
