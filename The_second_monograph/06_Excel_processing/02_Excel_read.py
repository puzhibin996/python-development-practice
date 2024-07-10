from openpyxl import load_workbook

workbook  = load_workbook('test.xlsx')
print(workbook.sheetnames)
sheet = workbook['Sheet3']
